from openpyxl import load_workbook

class ExcelWriter():
    
    def writeNode(self, data: dict, counter: int):

        if data.get('type') != 'node':
            return f"Datos no son de tipo 'node'. (type={data.get('type')})"
        

        row_value = data['value']  # Por ejemplo, 7, 3, etc.


        wb = load_workbook('test.xlsx')
        ws = wb['Day']

        # Ejemplo: escribe "Node" en alguna celda que dependa de row_value.
        row_index = counter + 1
        col_index = 2 
        ws.cell(row=col_index, column=col_index).value = "Node"
        ws.cell(row=row_index,column=col_index).value = row_value
        # Importante: guardar los cambios
        wb.save('test.xlsx')     

    def writeHumidity(self, data: dict, counter: int):
        if data.get('type') != 'humidity':
            return f"Datos no son de tipo 'humidity'. (type={data.get('type')})"
        
        row_value = data['value']      
        
        wb = load_workbook('test.xlsx')
        ws = wb['Day']

        # Ejemplo: escribe "Node" en alguna celda que dependa de row_value.
        row_index = counter
        col_index = 3 
        ws.cell(row=2, column=col_index).value = "Humidity"
        ws.cell(row=row_index,column=col_index).value = row_value
        # Importante: guardar los cambios
        wb.save('test.xlsx')        



